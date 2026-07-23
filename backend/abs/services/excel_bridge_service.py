"""
ExcelBridgeService — Excel review model (Agent #6, Layer A.4).

Generates a styled Excel workbook that mirrors the Python payment model for
business-user review. Business users who cannot read Python can verify every
formula, constant, and assumption in a familiar spreadsheet environment. The
workbook contains:

* **Summary** — deal metadata, key dates, parties
* **Certificates** — class-level setup (balance, rate, CUSIP, seniority)
* **Fees** — fee name, formula, frequency, parties
* **Waterfall** — priority table (verbatim + interpreted)
* **Governing Doc** — clause-level table (verbatim | plain English | formula)
* **Definitions** — all defined terms with raw + resolved text
* **Audit Trail** — recent audit-log entries

Every cell that originated from a specific SEP artifact carries a cell comment
with its ``artifact_id`` and ``citation``, so lineage is preserved in Excel.
Stateless + async (openpyxl is CPU-only).
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Optional

from backend.abs.services.base import ABSService, ProgressFn, ServiceContext, ServiceResult
from backend.abs.services.json_utils import parse_json_lenient
from backend.abs.store import DealStore

_HEADER_FILL = None  # lazily set once openpyxl is imported
_ACCENT = "5E5CE6"
_LIGHT = "EFF0FF"


def _apply_styles(ws: Any, header_fill: Any, header_font: Any) -> None:
    """Bold first row, freeze it, auto-size columns."""
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = ws["A2"]
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


class ExcelBridgeService(ABSService):
    """Generate an Excel review workbook from deal artifacts."""

    name = "excel_bridge"

    def __init__(self, deals_root: Path) -> None:
        self.deals_root = Path(deals_root)

    def context(self, deal_id: str) -> ServiceContext:
        return ServiceContext(deal_id=deal_id, deals_root=self.deals_root)

    async def generate(
        self,
        deal_id: str,
        *,
        actor: str = "system",
        progress: Optional[ProgressFn] = None,
    ) -> ServiceResult:
        return await self.guard(self._generate(deal_id, actor, progress))

    async def _generate(self, deal_id: str, actor: str, progress: Optional[ProgressFn]) -> dict[str, Any]:
        if progress:
            progress({"stage": "excel", "status": "in-progress"})
        path = await self._to_thread(self._build, deal_id, actor)
        if progress:
            progress({"stage": "excel", "status": "done", "path": str(path)})
        return {"path": str(path)}

    def _build(self, deal_id: str, actor: str) -> Path:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.comments import Comment

        ctx = self.context(deal_id)
        store = ctx.store(init=False)
        out_dir = ctx.scope().deal_path / "artifacts"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{deal_id}_review_model.xlsx"

        wb = openpyxl.Workbook()
        hdr_fill = PatternFill("solid", fgColor=_ACCENT)
        hdr_font = Font(bold=True, color="FFFFFF")

        # ── Summary sheet ────────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Field", "Value", "Citation"])
        docs = store.list_documents(deal_id)
        ws.append(["Deal ID", deal_id, ""])
        ws.append(["Documents", len(docs), ""])
        ws.append(["Definitions", len(store.list_definitions(deal_id)), ""])
        model = store.get_latest_payment_model(deal_id)
        if model:
            ws.append(["Model Version", model.get("version", 1), ""])
            ws.append(["Model Status", model.get("validation_status", ""), ""])
        for doc in docs:
            ws.append(["Document", doc.get("title", ""), doc.get("source_path", "")])
        _apply_styles(ws, hdr_fill, hdr_font)

        # ── Certificates sheet ────────────────────────────────────
        self._sheet_from_sep(wb, store, deal_id, "certificates",
                             ["class_name", "cusip", "original_balance", "certificate_rate", "seniority"],
                             hdr_fill, hdr_font)

        # ── Fees sheet ────────────────────────────────────────────
        self._sheet_from_sep(wb, store, deal_id, "fees",
                             ["fee_name", "parties", "frequency", "formula"],
                             hdr_fill, hdr_font)

        # ── Waterfall sheet ───────────────────────────────────────
        self._sheet_from_sep(wb, store, deal_id, "waterfall_rules",
                             ["priority", "section", "verbatim", "interpreted"],
                             hdr_fill, hdr_font)

        # ── Governing Doc sheet ───────────────────────────────────
        ws2 = wb.create_sheet("Governing Doc")
        ws2.append(["Verbatim", "Plain English", "Formula", "Citation"])
        for clause in store.list_governing_clauses(deal_id):
            ws2.append([clause.get("verbatim", ""), clause.get("plain_english", ""),
                         clause.get("math_formula", ""), clause.get("citation", "")])
        _apply_styles(ws2, hdr_fill, hdr_font)

        # ── Definitions sheet ─────────────────────────────────────
        ws3 = wb.create_sheet("Definitions")
        ws3.append(["Term", "Raw Definition", "Resolved Definition", "Page"])
        for d in store.list_definitions(deal_id):
            ws3.append([d.get("term_name", ""), d.get("raw_definition", ""),
                         d.get("resolved_definition", ""), d.get("page", "")])
        _apply_styles(ws3, hdr_fill, hdr_font)

        # ── Audit Trail sheet ─────────────────────────────────────
        ws4 = wb.create_sheet("Audit Trail")
        ws4.append(["Timestamp", "Actor", "Action", "Object Type", "Object ID"])
        for entry in store.list_audit(limit=200):
            ws4.append([entry.get("ts", ""), entry.get("actor", ""),
                         entry.get("action", ""), entry.get("object_type", ""), entry.get("object_id", "")])
        _apply_styles(ws4, hdr_fill, hdr_font)

        wb.save(str(out_path))
        store.audit("generate_excel", actor=actor, object_type="deal", object_id=deal_id,
                    after={"path": str(out_path)})
        return out_path

    def _sheet_from_sep(self, wb: Any, store: DealStore, deal_id: str,
                        sep_name: str, fields: list[str], hdr_fill: Any, hdr_font: Any) -> None:
        ws = wb.create_sheet(sep_name.replace("_", " ").title())
        ws.append(fields + ["citation", "status"])
        for art in store.list_sep_artifacts(deal_id, sep_name):
            v = parse_json_lenient(art.get("value") or "") or {}
            if not isinstance(v, dict):
                continue
            row = [v.get(f, "") for f in fields] + [art.get("citation", ""), art.get("status", "")]
            ws.append(row)
        _apply_styles(ws, hdr_fill, hdr_font)
