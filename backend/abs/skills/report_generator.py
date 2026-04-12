"""
Report Generator — Generate reports in Markdown, CSV, JSON, and Excel formats.
Supports governing docs, comparison reports, audit trails, and investor reports.

Ported from PayGen pipeline.skills.report_generator → backend.abs.skills
"""

from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_report(
    report_type: str,
    data: dict[str, Any],
    output_dir: Path,
    deal_id: str = "",
    format: str = "markdown",
) -> Path:
    """
    Generate a formatted report.

    Args:
        report_type: Type of report:
            - "governing_doc" → definitions/waterfall/accounts markdown
            - "comparison" → deal comparison report
            - "audit" → audit trail report
            - "cashflow" → cashflow projection summary
            - "quality" → quality/escalation report
        data: Report data dict
        output_dir: Output directory
        deal_id: Deal identifier
        format: Output format ("markdown", "csv", "json")

    Returns:
        Path to generated report file
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    generators = {
        "governing_doc": _generate_governing_doc,
        "comparison": _generate_comparison_report,
        "audit": _generate_audit_report,
        "cashflow": _generate_cashflow_report,
        "quality": _generate_quality_report,
    }

    generator = generators.get(report_type)
    if not generator:
        raise ValueError(f"Unknown report type: {report_type}. Valid: {list(generators.keys())}")

    content = generator(data, deal_id)

    # Write file
    ext_map = {"markdown": ".md", "csv": ".csv", "json": ".json", "excel": ".xlsx"}
    ext = ext_map.get(format, ".md")
    filename = f"{report_type}_{deal_id}_{timestamp}{ext}"
    output_path = output_dir / filename

    if format == "excel":
        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for Excel reports. "
                "Install with: pip install openpyxl"
            )
        _write_excel_report(output_path, report_type, data, deal_id, content)
    elif format == "json":
        output_path.write_text(
            json.dumps(data, indent=2, default=str),
            encoding="utf-8",
        )
    elif format == "csv" and isinstance(data.get("rows"), list):
        _write_csv(output_path, data["rows"], data.get("columns"))
    else:
        output_path.write_text(content, encoding="utf-8")

    return output_path


def generate_governing_docs(
    extractions: dict[str, list[dict]],
    output_dir: Path,
    deal_id: str = "",
) -> dict[str, Path]:
    """
    Generate full set of governing documents from extractions.

    Args:
        extractions: Dict mapping section names to extracted items
        output_dir: Output directory for governing docs
        deal_id: Deal identifier

    Returns:
        Dict mapping doc type to file path
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    generated: dict[str, Path] = {}

    # 1. Definitions
    defs = extractions.get("definitions", [])
    if defs:
        content = _format_definitions(defs, deal_id)
        path = output_dir / "01_definitions.md"
        path.write_text(content, encoding="utf-8")
        generated["definitions"] = path

    # 2. Waterfall Rules
    rules = extractions.get("waterfall_rules", [])
    if rules:
        content = _format_waterfall(rules, deal_id)
        path = output_dir / "02_waterfall.md"
        path.write_text(content, encoding="utf-8")
        generated["waterfall"] = path

    # 3. Accounts
    accts = extractions.get("accounts", [])
    if accts:
        content = _format_accounts(accts, deal_id)
        path = output_dir / "03_accounts.md"
        path.write_text(content, encoding="utf-8")
        generated["accounts"] = path

    # 4. Loss Allocations
    losses = extractions.get("loss_allocations", [])
    if losses:
        content = _format_loss_allocations(losses, deal_id)
        path = output_dir / "04_loss_allocation.md"
        path.write_text(content, encoding="utf-8")
        generated["loss_allocations"] = path

    # 5. Triggers
    triggers = extractions.get("triggers", [])
    if triggers:
        content = _format_triggers(triggers, deal_id)
        path = output_dir / "05_triggers.md"
        path.write_text(content, encoding="utf-8")
        generated["triggers"] = path

    return generated


# ── Report Generators ─────────────────────────────────────────

def _generate_governing_doc(data: dict, deal_id: str) -> str:
    """Generate a governing document section."""
    lines = [
        f"# Governing Document — {deal_id}",
        f"Generated: {datetime.now(timezone.utc).isoformat()}",
        "",
    ]

    for section, items in data.items():
        lines.append(f"## {section.replace('_', ' ').title()}")
        lines.append("")
        if isinstance(items, list):
            for item in items:
                if isinstance(item, dict):
                    item_id = item.get("id", item.get("name", ""))
                    lines.append(f"### {item_id}")
                    for k, v in item.items():
                        if k != "id":
                            lines.append(f"- **{k}**: {v}")
                    lines.append("")
                else:
                    lines.append(f"- {item}")
        lines.append("")

    return "\n".join(lines)


def _generate_comparison_report(data: dict, deal_id: str) -> str:
    """Generate a deal comparison report."""
    lines = [
        "# Deal Comparison Report",
        f"Generated: {datetime.now(timezone.utc).isoformat()}",
        "",
        f"**Deal A**: {data.get('deal_a_id', 'N/A')}",
        f"**Deal B**: {data.get('deal_b_id', 'N/A')}",
        f"**Overall Similarity**: {data.get('overall_similarity', 0):.1%}",
        "",
        "## Section Comparisons",
        "",
    ]

    sections = data.get("sections", data.get("section_comparisons", {}))
    for section_name, section_data in sections.items():
        sim = section_data.get("similarity_score", 0)
        lines.append(f"### {section_name}")
        lines.append(f"- Similarity: {sim:.1%}")
        lines.append(f"- Deal A items: {section_data.get('deal_a_count', 0)}")
        lines.append(f"- Deal B items: {section_data.get('deal_b_count', 0)}")
        lines.append(f"- Matched: {section_data.get('matched', 0)}")

        only_a = section_data.get("only_in_a", [])
        only_b = section_data.get("only_in_b", [])
        if only_a:
            lines.append(f"- Only in A: {', '.join(str(x) for x in only_a)}")
        if only_b:
            lines.append(f"- Only in B: {', '.join(str(x) for x in only_b)}")
        lines.append("")

    return "\n".join(lines)


def _generate_audit_report(data: dict, deal_id: str) -> str:
    """Generate an audit trail report."""
    lines = [
        f"# Audit Trail — {deal_id}",
        f"Generated: {datetime.now(timezone.utc).isoformat()}",
        "",
    ]

    entries = data.get("entries", data.get("audit_trail", []))
    for entry in entries:
        ts = entry.get("timestamp", "")
        agent = entry.get("agent", "")
        action = entry.get("action", "")
        details = entry.get("details", "")
        lines.append(f"- **[{ts}]** `{agent}`: {action}")
        if details:
            lines.append(f"  - {details}")

    return "\n".join(lines)


def _generate_cashflow_report(data: dict, deal_id: str) -> str:
    """Generate cashflow projection summary."""
    lines = [
        f"# Cashflow Projection — {deal_id}",
        f"Scenario: {data.get('scenario', 'base')}",
        f"Generated: {datetime.now(timezone.utc).isoformat()}",
        "",
    ]

    summary = data.get("summary", {})
    lines.append("## Summary")
    lines.append(f"- Total months: {summary.get('total_months', 0)}")
    lines.append(f"- Total collections: ${summary.get('total_collections', 0):,.2f}")
    lines.append(f"- Total distributions: ${summary.get('total_distributions', 0):,.2f}")
    lines.append("")

    final = summary.get("final_class_balances", {})
    if final:
        lines.append("## Final Balances")
        for cls, balance in final.items():
            lines.append(f"- {cls}: ${balance:,.2f}")

    return "\n".join(lines)


def _generate_quality_report(data: dict, deal_id: str) -> str:
    """Generate quality and escalation report."""
    lines = [
        f"# Quality Report — {deal_id}",
        f"Generated: {datetime.now(timezone.utc).isoformat()}",
        "",
    ]

    # Quality scores
    scores = data.get("quality_scores", {})
    if scores:
        lines.append("## Quality Scores")
        for dim, score in scores.items():
            lines.append(f"- {dim}: {score}/10")
        lines.append("")

    # Escalations
    escalations = data.get("escalations", [])
    if escalations:
        lines.append("## Escalations")
        for esc in escalations:
            severity = esc.get("severity", "unknown")
            desc = esc.get("description", "")
            lines.append(f"- [{severity.upper()}] {desc}")

    return "\n".join(lines)


# ── Governing Doc Formatters ──────────────────────────────────

def _format_definitions(defs: list[dict], deal_id: str) -> str:
    """Format definitions into a governing document."""
    lines = [
        f"# Definitions — {deal_id}",
        "",
        "## Key Terms and Definitions",
        "",
    ]
    for d in defs:
        term = d.get("term", d.get("id", "Unknown"))
        definition = d.get("definition", d.get("description", ""))
        source = d.get("source_section", "")
        lines.append(f"### {term}")
        lines.append(f"{definition}")
        if source:
            lines.append(f"*Source: {source}*")
        lines.append("")
    return "\n".join(lines)


def _format_waterfall(rules: list[dict], deal_id: str) -> str:
    """Format waterfall rules."""
    lines = [
        f"# Payment Waterfall — {deal_id}",
        "",
        "## Distribution Rules (Priority Order)",
        "",
    ]
    for i, rule in enumerate(rules, 1):
        step = rule.get("step", rule.get("id", f"Step {i}"))
        desc = rule.get("description", "")
        target = rule.get("target", "")
        lines.append(f"### {i}. {step}")
        if target:
            lines.append(f"**Target**: {target}")
        lines.append(f"{desc}")
        lines.append("")
    return "\n".join(lines)


def _format_accounts(accts: list[dict], deal_id: str) -> str:
    """Format accounts."""
    lines = [
        f"# Trust Accounts — {deal_id}",
        "",
    ]
    for acct in accts:
        name = acct.get("name", acct.get("id", ""))
        purpose = acct.get("purpose", acct.get("description", ""))
        lines.append(f"### {name}")
        lines.append(f"{purpose}")
        lines.append("")
    return "\n".join(lines)


def _format_loss_allocations(losses: list[dict], deal_id: str) -> str:
    """Format loss allocation rules."""
    lines = [
        f"# Loss Allocation — {deal_id}",
        "",
    ]
    for loss in losses:
        name = loss.get("name", loss.get("id", ""))
        desc = loss.get("description", loss.get("methodology", ""))
        lines.append(f"### {name}")
        lines.append(f"{desc}")
        lines.append("")
    return "\n".join(lines)


def _format_triggers(triggers: list[dict], deal_id: str) -> str:
    """Format trigger definitions."""
    lines = [
        f"# Performance Triggers — {deal_id}",
        "",
    ]
    for trig in triggers:
        name = trig.get("name", trig.get("id", ""))
        condition = trig.get("condition", trig.get("description", ""))
        effect = trig.get("effect", "")
        lines.append(f"### {name}")
        lines.append(f"**Condition**: {condition}")
        if effect:
            lines.append(f"**Effect**: {effect}")
        lines.append("")
    return "\n".join(lines)


# ── Helpers ───────────────────────────────────────────────────

def _write_csv(path: Path, rows: list[dict], columns: Optional[list[str]] = None) -> None:
    """Write rows to CSV."""
    if not rows:
        path.write_text("", encoding="utf-8")
        return

    if columns is None:
        columns = list(rows[0].keys())

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


# ── Excel Report Generation ──────────────────────────────────

def generate_excel_report(
    data: dict[str, Any],
    output_path: Path,
    deal_id: str = "",
    report_title: str = "Investor Report",
) -> Path:
    """
    Generate a formatted Excel (.xlsx) investor report.

    Args:
        data: Report data dict. Expected keys:
            - "summary": dict with deal-level metrics
            - "classes": list of dicts with per-class data
            - "cashflows": list of dicts with monthly cashflow data
            - "triggers": list of dicts with trigger status
            - Any additional keys become separate worksheets
        output_path: Path for the output .xlsx file
        deal_id: Deal identifier
        report_title: Report title for the cover sheet

    Returns:
        Path to generated Excel file

    Raises:
        ImportError: If openpyxl is not installed
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel reports. "
            "Install with: pip install openpyxl"
        )

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_font = Font(name="Calibri", size=16, bold=True, color="2F5496")
    currency_format = '#,##0.00'
    pct_format = '0.00%'
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── Cover Sheet ──
    ws_cover = wb.active
    ws_cover.title = "Summary"
    ws_cover["A1"] = report_title
    ws_cover["A1"].font = title_font
    ws_cover["A2"] = f"Deal: {deal_id}"
    ws_cover["A3"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws_cover["A4"] = ""

    # Summary data
    summary = data.get("summary", {})
    row = 5
    for key, value in summary.items():
        ws_cover.cell(row=row, column=1, value=key.replace("_", " ").title())
        ws_cover.cell(row=row, column=1).font = Font(bold=True)
        cell = ws_cover.cell(row=row, column=2, value=value)
        if isinstance(value, float) and abs(value) >= 1:
            cell.number_format = currency_format
        elif isinstance(value, float) and abs(value) < 1:
            cell.number_format = pct_format
        row += 1

    ws_cover.column_dimensions["A"].width = 30
    ws_cover.column_dimensions["B"].width = 25

    # ── Classes Sheet ──
    classes = data.get("classes", [])
    if classes:
        ws_classes = wb.create_sheet("Classes")
        _write_excel_table(
            ws_classes, classes,
            header_font, header_fill, header_alignment,
            thin_border, currency_format,
        )

    # ── Cashflows Sheet ──
    cashflows = data.get("cashflows", data.get("monthly_data", []))
    if cashflows:
        ws_cf = wb.create_sheet("Cashflows")
        _write_excel_table(
            ws_cf, cashflows,
            header_font, header_fill, header_alignment,
            thin_border, currency_format,
        )

    # ── Triggers Sheet ──
    triggers = data.get("triggers", [])
    if triggers:
        ws_trig = wb.create_sheet("Triggers")
        _write_excel_table(
            ws_trig, triggers,
            header_font, header_fill, header_alignment,
            thin_border, None,
        )

    # ── Additional sheets for any other keys ──
    standard_keys = {"summary", "classes", "cashflows", "monthly_data",
                     "triggers", "report_title"}
    for key, value in data.items():
        if key in standard_keys:
            continue
        if isinstance(value, list) and value and isinstance(value[0], dict):
            ws = wb.create_sheet(key.replace("_", " ").title()[:31])
            _write_excel_table(
                ws, value,
                header_font, header_fill, header_alignment,
                thin_border, currency_format,
            )

    wb.save(str(output_path))
    return output_path


def _write_excel_table(
    ws: Any,
    rows: list[dict],
    header_font: Any,
    header_fill: Any,
    header_alignment: Any,
    border: Any,
    number_format: Optional[str],
) -> None:
    """Write a list of dicts as a formatted table to a worksheet."""
    if not rows:
        return

    columns = list(rows[0].keys())

    # Headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        if border:
            cell.border = border

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if border:
                cell.border = border
            if isinstance(value, float) and number_format:
                cell.number_format = number_format

    # Auto-fit column widths (approximate)
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row_idx in range(2, len(rows) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(val))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"


def _write_excel_report(
    output_path: Path,
    report_type: str,
    data: dict,
    deal_id: str,
    markdown_content: str,
) -> None:
    """Write a report as an Excel file with the content on the first sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.replace("_", " ").title()

    # Write markdown content as plain text rows
    for row_idx, line in enumerate(markdown_content.split("\n"), 1):
        ws.cell(row=row_idx, column=1, value=line)

    # If data has tabular content, add as a second sheet
    for key, value in data.items():
        if isinstance(value, list) and value and isinstance(value[0], dict):
            ws2 = wb.create_sheet(key.replace("_", " ").title()[:31])
            columns = list(value[0].keys())
            for col_idx, col_name in enumerate(columns, 1):
                ws2.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row_data in enumerate(value, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    ws2.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

    ws.column_dimensions["A"].width = 100
    wb.save(str(output_path))
