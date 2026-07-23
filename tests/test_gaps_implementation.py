"""
Tests for all newly implemented gaps:
 - SEP-7 auto-discover + SEP-8 gap_review profiles
 - DealSetupService (Layer A.3): tblCertInfo, fee, account, trigger, reporting CSVs + manifest
 - ProductionReadinessGate (monthly input validation)
 - ModelService.generate_spec (model-spec intermediate)
 - ModelService self-heal loop
 - ExcelBridgeService (Excel review workbook)
 - EvidencePackageService (FRD Screen 8 bundle)
 - SourceHierarchyService (Layer B.2: conflict detection + hierarchy)
 - CommandCenterService (Layer B.9: cross-deal ops queue)
 - New dispatcher commands: setup.generate, evidence.generate, excel.generate,
   model.spec, model.readiness, hierarchy.*, command_center.queue
"""

from __future__ import annotations

import asyncio
import json
from pathlib import Path

import pytest

from backend.abs.services import IngestionService, SEPService, StubLLMClient
from backend.abs.services.command_center_service import CommandCenterService
from backend.abs.services.deal_setup_service import DealSetupService
from backend.abs.services.deal_service import DealService
from backend.abs.services.dispatcher import ABSDispatcher
from backend.abs.services.evidence_service import EvidencePackageService
from backend.abs.services.excel_bridge_service import ExcelBridgeService
from backend.abs.services.model_run_service import ProductionReadinessGate
from backend.abs.services.model_service import ModelService
from backend.abs.services.pdf_extract import extracted_from_pages
from backend.abs.services.sep_profiles import AUTO_DISCOVER, CORE_PROFILES, GAP_REVIEW, TRIGGERS, DEAL_PARTIES, get_profile
from backend.abs.services.source_hierarchy_service import SourceHierarchyService
from backend.abs.store import DealStore

PAGES = [
    "ARTICLE I DEFINITIONS\n\nAvailable Funds means the Net Interest plus principal.\n\n",
    "ARTICLE V FEES\n\nThe Servicing Fee equals 0.50% per annum payable monthly to Servicer.\n\n",
    "ARTICLE VI CERTIFICATES\n\nClass A-1 CUSIP 12489WEX8 original balance 90,650,000 rate 5.02956% senior.\n\n",
]


def _ingest_and_seed(tmp_path: Path) -> DealStore:
    doc = extracted_from_pages(PAGES)
    asyncio.run(IngestionService(tmp_path).ingest_document("cbass", extracted=doc, doc_type="PSA"))
    store = DealStore.for_deal_dir(tmp_path / "cbass", init=False)
    # Seed approved certificate + fee artifacts
    store.add_sep_artifact({"deal_id": "cbass", "sep_name": "certificates",
                            "value": {"class_name": "A-1", "cusip": "12489WEX8",
                                      "original_balance": "90,650,000.00", "accrual_formula": "5.02956%",
                                      "seniority": "senior"}, "citation": "Art VI p.3", "status": "approved"})
    store.add_sep_artifact({"deal_id": "cbass", "sep_name": "fees",
                            "value": {"fee_name": "Servicing Fee", "parties": "Servicer",
                                      "frequency": "monthly", "formula": "0.50% * pool_balance / 12"},
                            "citation": "Art V p.2", "status": "approved"})
    store.add_sep_artifact({"deal_id": "cbass", "sep_name": "accounts",
                            "value": {"account_name": "Collection Account", "account_type": "trust",
                                      "inflows": "principal + interest", "outflows": "distributions"},
                            "citation": "Art III p.2", "status": "approved"})
    return store


# ---------------------------------------------------------------------------
# New SEP profiles
# ---------------------------------------------------------------------------

def test_new_sep_profiles_in_catalog():
    names = {p.name for p in CORE_PROFILES}
    assert {"deal_parties", "triggers", "auto_discover", "gap_review"} <= names


def test_auto_discover_runs_via_dispatcher(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    d = ABSDispatcher(tmp_path)
    llm = StubLLMClient(responder=lambda p, s: json.dumps([
        {"proposed_sep": "Clean-up Call Duties", "duty_category": "cert_admin",
         "responsible_party": "Certificate Administrator", "rationale": "Found in Art VII",
         "citation": "Art VII p.5"}
    ]))
    res = asyncio.run(d.dispatch("sep.run", {"deal_id": "cbass", "sep_name": "auto_discover"}, llm=llm))
    assert res["ok"] is True
    assert res["data"]["items"] == 1


# ---------------------------------------------------------------------------
# DealSetupService (Layer A.3)
# ---------------------------------------------------------------------------

def test_deal_setup_generates_certinfo_csv(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    svc = DealSetupService(tmp_path)
    events: list[dict] = []
    res = asyncio.run(svc.generate("cbass", progress=events.append))
    assert res.ok is True, res.error
    files = res.data["files"]
    assert "cert_info_csv" in files
    cert_csv = Path(files["cert_info_csv"])
    assert cert_csv.exists()
    content = cert_csv.read_text()
    assert "A-1" in content and "12489WEX8" in content and "90650000.0" in content
    assert "fee_setup_csv" in files
    assert "account_setup_csv" in files
    assert "manifest" in files
    assert "validation_report" in files
    assert {"in-progress", "done"} == {e["status"] for e in events}


def test_deal_setup_manifest_has_lineage(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    svc = DealSetupService(tmp_path)
    asyncio.run(svc.generate("cbass"))
    manifest_path = tmp_path / "cbass" / "artifacts" / "setup" / "setup_manifest.json"
    assert manifest_path.exists()
    manifest = json.loads(manifest_path.read_text())
    assert manifest["deal_id"] == "cbass"
    assert len(manifest["artifacts"]) >= 3


def test_deal_setup_validation_flags_missing_cusip(tmp_path: Path):
    store = DealStore.for_deal_dir(tmp_path / "cbass")
    store.add_sep_artifact({"deal_id": "cbass", "sep_name": "certificates",
                            "value": {"class_name": "B-1", "original_balance": "5000000"},  # no CUSIP
                            "citation": "p.5", "status": "approved"})
    svc = DealSetupService(tmp_path)
    res = asyncio.run(svc.generate("cbass"))
    assert res.ok is True
    issues = res.data["validation_issues"]
    assert any(i.get("field") == "cusip" and i.get("severity") == "warning" for i in issues)


def test_setup_generate_via_dispatcher(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    d = ABSDispatcher(tmp_path)
    res = asyncio.run(d.dispatch("setup.generate", {"deal_id": "cbass"}))
    assert res["ok"] is True
    assert "cert_info_csv" in res["data"]["files"]


# ---------------------------------------------------------------------------
# Production readiness gate
# ---------------------------------------------------------------------------

def test_readiness_gate_passes_valid_inputs():
    gate = ProductionReadinessGate()
    result = gate.check([{"interest_collections": 10000.0, "principal_collections": 50000.0}])
    assert result["ready"] is True
    assert result["errors"] == 0


def test_readiness_gate_fails_non_numeric():
    gate = ProductionReadinessGate()
    result = gate.check([{"interest_collections": "N/A", "principal_collections": 50000.0}])
    assert result["ready"] is False
    assert result["errors"] >= 1


def test_readiness_gate_warns_missing_field():
    gate = ProductionReadinessGate()
    result = gate.check([{"interest_collections": 100.0}])  # missing principal_collections
    assert result["warnings"] >= 1


def test_model_run_blocked_by_readiness(tmp_path: Path):
    DealStore.for_deal_dir(tmp_path / "cbass")
    from backend.abs.services.model_run_service import ModelRunService

    svc = ModelRunService(tmp_path)
    bad_inputs = [{"interest_collections": "BAD", "principal_collections": 0.0}]
    res = asyncio.run(svc.run("cbass", monthly_inputs=bad_inputs))
    assert res.ok is False
    assert "readiness" in res.error.lower()


def test_model_readiness_command(tmp_path: Path):
    d = ABSDispatcher(tmp_path)
    res = asyncio.run(d.dispatch("model.readiness", {
        "monthly_inputs": [{"interest_collections": 100.0, "principal_collections": 50.0}]
    }))
    assert res["ok"] is True
    assert res["data"]["ready"] is True


# ---------------------------------------------------------------------------
# ModelService.generate_spec
# ---------------------------------------------------------------------------

def test_model_spec_creates_file(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    store = DealStore.for_deal_dir(tmp_path / "cbass", init=False)
    store.add_governing_clause({"deal_id": "cbass", "verbatim": "pay A-1 first",
                                "plain_english": "Class A-1 gets interest then principal",
                                "math_formula": "rate/12*bal", "citation": "Art V p.2"})
    svc = ModelService(tmp_path)
    llm = StubLLMClient(responder=lambda p, s: "MODEL SPEC:\n1. Pay Class A-1 interest first [Art V p.2]\n2. Then principal.")
    res = asyncio.run(svc.generate_spec("cbass", llm))
    assert res.ok is True
    assert "spec" in res.data
    assert Path(res.data["spec_path"]).exists()
    assert "MODEL SPEC" in Path(res.data["spec_path"]).read_text()


def test_model_spec_via_dispatcher(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    d = ABSDispatcher(tmp_path)
    llm = StubLLMClient(responder=lambda p, s: "MODEL SPEC: waterfall order as per Art V.")
    res = asyncio.run(d.dispatch("model.spec", {"deal_id": "cbass"}, llm=llm))
    assert res["ok"] is True
    assert res["data"]["spec"]


# ---------------------------------------------------------------------------
# ModelService self-heal loop
# ---------------------------------------------------------------------------

def test_model_generate_self_heals_on_audit_fail(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    store = DealStore.for_deal_dir(tmp_path / "cbass", init=False)
    store.add_governing_clause({"deal_id": "cbass", "verbatim": "pay A-1",
                                "plain_english": "Pay Class A-1", "math_formula": "rate/12*bal",
                                "citation": "p.2"})
    call_count = [0]

    def responder(prompt: str, system: str) -> str:
        call_count[0] += 1
        if "auditor" in system.lower():
            if call_count[0] <= 2:  # fail first 2 audits to trigger heal
                return json.dumps({"checks": [{"item": "A-1 rate", "pass": False, "source": "", "note": "missing cite"}], "verdict": "fail"})
            return json.dumps({"checks": [{"item": "A-1 rate", "pass": True, "source": "p.2", "note": "ok"}], "verdict": "pass"})
        # code gen or self-heal
        return "class WaterfallModel:\n    def run_month(self, inputs):  # cite: p.2\n        return {}\n"

    svc = ModelService(tmp_path)
    res = asyncio.run(svc.generate("cbass", StubLLMClient(responder=responder)))
    assert res.ok is True
    assert res.data["heal_iters"] >= 1  # at least one heal iteration happened


# ---------------------------------------------------------------------------
# ExcelBridgeService
# ---------------------------------------------------------------------------

def test_excel_bridge_generates_workbook(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    svc = ExcelBridgeService(tmp_path)
    events: list[dict] = []
    res = asyncio.run(svc.generate("cbass", progress=events.append))
    assert res.ok is True, res.error
    path = Path(res.data["path"])
    assert path.exists() and path.suffix == ".xlsx"
    import openpyxl
    wb = openpyxl.load_workbook(str(path))
    assert "Summary" in wb.sheetnames
    assert "Certificates" in wb.sheetnames
    assert "Fees" in wb.sheetnames
    ws = wb["Certificates"]
    cell_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert "A-1" in cell_values


def test_excel_via_dispatcher(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    d = ABSDispatcher(tmp_path)
    res = asyncio.run(d.dispatch("excel.generate", {"deal_id": "cbass"}))
    assert res["ok"] is True
    assert Path(res["data"]["path"]).suffix == ".xlsx"


# ---------------------------------------------------------------------------
# EvidencePackageService (FRD Screen 8)
# ---------------------------------------------------------------------------

def test_evidence_package_generated(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    svc = EvidencePackageService(tmp_path)
    events: list[dict] = []
    res = asyncio.run(svc.generate("cbass", actor="auditor", progress=events.append))
    assert res.ok is True, res.error
    json_path = Path(res.data["json_path"])
    summary_path = Path(res.data["summary_path"])
    assert json_path.exists() and summary_path.exists()
    pkg = json.loads(json_path.read_text())
    assert pkg["deal_id"] == "cbass"
    assert "source_documents" in pkg
    assert "approval_history" in pkg
    assert pkg["schema_version"] >= 3
    summary = summary_path.read_text()
    assert "EVIDENCE PACKAGE" in summary


def test_evidence_via_dispatcher(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    d = ABSDispatcher(tmp_path)
    res = asyncio.run(d.dispatch("evidence.generate", {"deal_id": "cbass"}))
    assert res["ok"] is True
    assert Path(res["data"]["json_path"]).exists()


# ---------------------------------------------------------------------------
# SourceHierarchyService (Layer B.2)
# ---------------------------------------------------------------------------

def test_hierarchy_no_conflict_single_doc(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    svc = SourceHierarchyService(tmp_path)
    res = asyncio.run(svc.detect_conflicts("cbass"))
    assert res.ok is True
    assert res.data == []  # only one document → no cross-document conflicts


def test_hierarchy_detects_multi_doc_conflict(tmp_path: Path):
    store = DealStore.for_deal_dir(tmp_path / "cbass")
    doc_a = store.add_document({"deal_id": "cbass", "doc_type": "PSA", "title": "PSA"})
    doc_b = store.add_document({"deal_id": "cbass", "doc_type": "Amendment", "title": "Amend 1"})
    store.add_sep_artifact({"deal_id": "cbass", "sep_name": "fees",
                            "field_path": "servicer_fee", "value": "0.005", "citation": "PSA p.30",
                            "status": "approved"})
    store.add_sep_artifact({"deal_id": "cbass", "sep_name": "fees",
                            "field_path": "servicer_fee", "value": "0.0055", "citation": "Amendment p.2",
                            "status": "approved"})
    svc = SourceHierarchyService(tmp_path)
    res = asyncio.run(svc.detect_conflicts("cbass"))
    assert res.ok is True
    assert any(c["field_path"] == "servicer_fee" for c in res.data)


def test_hierarchy_get_and_confirm(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    store = DealStore.for_deal_dir(tmp_path / "cbass", init=False)
    docs = store.list_documents("cbass")
    svc = SourceHierarchyService(tmp_path)
    h_res = asyncio.run(svc.get_hierarchy("cbass"))
    assert h_res.ok is True
    assert len(h_res.data) >= 1
    confirm = asyncio.run(svc.confirm_operative("cbass", docs[0]["doc_id"],
                                                 logic_area="waterfall", actor="reviewer"))
    assert confirm.ok is True
    assert confirm.data["confirmed"] is True


def test_hierarchy_via_dispatcher(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    d = ABSDispatcher(tmp_path)
    res = asyncio.run(d.dispatch("hierarchy.detect", {"deal_id": "cbass"}))
    assert res["ok"] is True


# ---------------------------------------------------------------------------
# CommandCenterService (Layer B.9)
# ---------------------------------------------------------------------------

def test_command_center_shows_pending_reviews(tmp_path: Path):
    _ingest_and_seed(tmp_path)
    # Add an artifact that's still pending_review
    store = DealStore.for_deal_dir(tmp_path / "cbass", init=False)
    store.add_sep_artifact({"deal_id": "cbass", "sep_name": "accounts",
                            "value": {"account_name": "Pending Account"}, "status": "pending_review"})
    svc = CommandCenterService(tmp_path)
    res = asyncio.run(svc.queue())
    assert res.ok is True
    item_types = {i["type"] for i in res.data["items"]}
    assert "pending_review" in item_types


def test_command_center_multi_deal(tmp_path: Path):
    deal_svc = DealService(tmp_path)
    asyncio.run(deal_svc.create_deal("deal_a"))
    asyncio.run(deal_svc.create_deal("deal_b"))
    store_a = DealStore.for_deal_dir(tmp_path / "deal_a", init=False)
    store_a.add_sep_artifact({"deal_id": "deal_a", "sep_name": "fees",
                              "value": {"fee_name": "Fee A"}, "status": "pending_review"})
    store_b = DealStore.for_deal_dir(tmp_path / "deal_b", init=False)
    store_b.add_correction_event({"deal_id": "deal_b", "severity": "high", "root_cause": "bad extract"})
    svc = CommandCenterService(tmp_path)
    res = asyncio.run(svc.queue())
    assert res.ok is True
    deal_ids = {i["deal_id"] for i in res.data["items"]}
    assert {"deal_a", "deal_b"} <= deal_ids


def test_command_center_via_dispatcher(tmp_path: Path):
    asyncio.run(DealService(tmp_path).create_deal("cbass"))
    d = ABSDispatcher(tmp_path)
    res = asyncio.run(d.dispatch("command_center.queue", {}))
    assert res["ok"] is True
    assert "items" in res["data"]
